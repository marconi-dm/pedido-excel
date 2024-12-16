from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO
import pandas as pd


class SpreadsheetBuilder:
    
    @staticmethod
    def save_to_excel(dataframe):
        """
        Salva o DataFrame em um arquivo Excel na memória, aplicando formatações básicas.

        Parâmetros:
        dataframe (DataFrame): O DataFrame pandas com os dados estruturados.

        Retorno:
        BytesIO: Um buffer contendo o arquivo Excel.
        """
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        dataframe.to_excel(writer, index=False, sheet_name='Processed Data')
        worksheet = writer.sheets['Processed Data']

        # Formatação do cabeçalho
        header_fill = PatternFill(start_color='72BF78', end_color='72BF78', fill_type='solid')
        header_font = Font(size=12, bold=True, color='FFFFFF')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Cores alternadas para as linhas
        row_fill1 = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
        row_fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        thin_border = Side(border_style='thin', color='000000')
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        toggle = True  # Alterna entre as cores para cada linha
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.fill = row_fill1 if toggle else row_fill2
                cell.border = border
            toggle = not toggle

        writer.close()
        output.seek(0)  # Volta o ponteiro do buffer para o início
        return output


